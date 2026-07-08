import csv
import json
import os
import zipfile
from itertools import chain
from pathlib import Path
from uuid import uuid4

import pandas as pd
from fastapi import HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from app.models.dataset import Dataset
from app.models.user import User


UPLOAD_DIR = Path(__file__).resolve().parents[2] / "uploads"
MAX_UPLOAD_SIZE_BYTES = int(os.getenv("MAX_UPLOAD_SIZE_MB", "50")) * 1024 * 1024
MAX_DATAFRAME_MEMORY_BYTES = (
    int(os.getenv("MAX_DATAFRAME_MEMORY_MB", "200")) * 1024 * 1024
)
UPLOAD_CHUNK_SIZE_BYTES = 1024 * 1024
CSV_CHUNK_SIZE_ROWS = 10_000
CSV_CONCATENATION_MEMORY_FACTOR = 2


class DatasetReadError(Exception):
    """Raised when a dataset cannot be parsed or read from storage."""


class DatasetMemoryLimitError(DatasetReadError):
    """Raised when reading a dataset would exceed the configured memory limit."""


def ensure_unique_header_values(header_values: list[object]) -> None:
    normalized_headers = [str(value) for value in header_values]

    if len(normalized_headers) != len(set(normalized_headers)):
        raise DatasetReadError(
            "Dataset contains duplicate or ambiguous column names"
        )


def read_csv_header(file_path: str) -> list[object]:
    with open(file_path, encoding="utf-8-sig", newline="") as dataset_file:
        for line in dataset_file:
            if line.strip():
                return next(csv.reader(chain([line], dataset_file)), [])

    return []


def read_xlsx_header(file_path: str) -> list[object]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            return []
        worksheet = workbook.worksheets[0]
        return list(
            next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
        )
    finally:
        workbook.close()


def read_xls_header(file_path: str) -> list[object]:
    import xlrd

    workbook = xlrd.open_workbook(file_path, on_demand=True)
    try:
        if workbook.nsheets == 0:
            return []
        worksheet = workbook.sheet_by_index(0)
        return worksheet.row_values(0) if worksheet.nrows else []
    finally:
        workbook.release_resources()


def validate_raw_dataset_headers(file_path: str, file_extension: str) -> None:
    if file_extension == ".csv":
        header_values = read_csv_header(file_path)
    elif file_extension == ".xlsx":
        header_values = read_xlsx_header(file_path)
    elif file_extension == ".xls":
        header_values = read_xls_header(file_path)
    else:
        raise DatasetReadError("Unsupported dataset file format")

    ensure_unique_header_values(header_values)


def normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized_columns = [str(column) for column in df.columns]

    if len(normalized_columns) != len(set(normalized_columns)):
        raise DatasetReadError(
            "Dataset contains ambiguous column names after converting them to text"
        )

    df.columns = normalized_columns
    return df


def get_dataframe_memory_usage(df: pd.DataFrame) -> int:
    return int(df.memory_usage(index=True, deep=True).sum())


def ensure_dataframe_fits_memory_limit(memory_usage_bytes: int) -> None:
    if memory_usage_bytes > MAX_DATAFRAME_MEMORY_BYTES:
        raise DatasetMemoryLimitError(
            "Dataset requires more memory than the configured limit"
        )


def read_csv_with_memory_limit(file_path: str) -> pd.DataFrame:
    chunks = []
    memory_usage_bytes = 0

    for chunk in pd.read_csv(file_path, chunksize=CSV_CHUNK_SIZE_ROWS):
        memory_usage_bytes += get_dataframe_memory_usage(chunk)
        ensure_dataframe_fits_memory_limit(
            memory_usage_bytes * CSV_CONCATENATION_MEMORY_FACTOR
        )
        chunks.append(chunk)

    if not chunks:
        return pd.read_csv(file_path)

    dataframe = pd.concat(chunks, ignore_index=True)
    ensure_dataframe_fits_memory_limit(get_dataframe_memory_usage(dataframe))
    return dataframe


def validate_excel_expanded_size(file_path: str) -> None:
    if not zipfile.is_zipfile(file_path):
        return

    with zipfile.ZipFile(file_path) as workbook:
        expanded_size = sum(entry.file_size for entry in workbook.infolist())

    ensure_dataframe_fits_memory_limit(expanded_size)


def read_dataset_file(file_path: str) -> pd.DataFrame:
    file_extension = Path(file_path).suffix.lower()

    try:
        if file_extension == ".xlsx":
            validate_excel_expanded_size(file_path)

        validate_raw_dataset_headers(file_path, file_extension)

        if file_extension == ".csv":
            return normalize_dataframe_columns(read_csv_with_memory_limit(file_path))

        if file_extension in {".xlsx", ".xls"}:
            df = pd.read_excel(file_path)
            ensure_dataframe_fits_memory_limit(get_dataframe_memory_usage(df))
            return normalize_dataframe_columns(df)

        raise DatasetReadError("Unsupported dataset file format")

    except DatasetReadError:
        raise

    except Exception as exc:
        raise DatasetReadError("Dataset file could not be read") from exc


def read_stored_dataset_file(file_path: str) -> pd.DataFrame:
    try:
        return read_dataset_file(file_path)
    except DatasetReadError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Stored dataset could not be read",
        ) from exc


def get_owned_dataset(
    db: Session,
    dataset_id: int,
    current_user: User,
) -> Dataset:
    dataset = (
        db.query(Dataset)
        .filter(
            Dataset.id == dataset_id,
            Dataset.user_id == current_user.id,
        )
        .first()
    )

    if dataset is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset not found",
        )

    return dataset


def upload_dataset(db: Session, file: UploadFile, current_user: User) -> Dataset:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file uploaded",
        )

    allowed_extensions = {".csv", ".xlsx", ".xls"}

    file_extension = Path(file.filename).suffix.lower()

    if file_extension not in allowed_extensions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files are allowed",
        )

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    unique_filename = f"{uuid4()}_{Path(file.filename).name}"
    file_path = UPLOAD_DIR / unique_filename

    try:
        with open(file_path, "wb") as buffer:
            uploaded_bytes = 0
            while chunk := file.file.read(UPLOAD_CHUNK_SIZE_BYTES):
                uploaded_bytes += len(chunk)
                if uploaded_bytes > MAX_UPLOAD_SIZE_BYTES:
                    raise HTTPException(
                        status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                        detail="Uploaded file exceeds the maximum allowed size",
                    )
                buffer.write(chunk)

        try:
            df = read_dataset_file(str(file_path))
        except DatasetMemoryLimitError as exc:
            raise HTTPException(
                status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                detail="Dataset requires more memory than the allowed limit",
            ) from exc
        except DatasetReadError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file could not be read as a valid CSV or Excel file",
            ) from exc

        row_count, column_count = df.shape

        new_dataset = Dataset(
            user_id=current_user.id,
            file_name=file.filename,
            file_path=str(file_path),
            row_count=row_count,
            column_count=column_count,
        )

        db.add(new_dataset)
        db.commit()
        db.refresh(new_dataset)

        return new_dataset

    except HTTPException:
        file_path.unlink(missing_ok=True)
        raise
    except Exception:
        db.rollback()
        file_path.unlink(missing_ok=True)
        raise
    finally:
        file.file.close()


def get_dataset_preview(
    db: Session,
    dataset_id: int,
    current_user: User,
) -> dict:
    dataset = get_owned_dataset(db, dataset_id, current_user)

    if not Path(dataset.file_path).is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset file not found on server",
        )

    df = read_stored_dataset_file(dataset.file_path)

    preview_df = df.head(10)

    column_info = []
    for column in df.columns:
        missing_count = int(df[column].isna().sum())
        missing_percentage = (
            round((missing_count / len(df)) * 100, 2) if len(df) > 0 else 0
        )

        column_info.append(
            {
                "name": str(column),
                "dtype": str(df[column].dtype),
                "missing_count": missing_count,
                "missing_percentage": missing_percentage,
            }
        )

    duplicate_rows = int(df.duplicated().sum())

    preview = json.loads(preview_df.to_json(orient="records", date_format="iso"))
    summary_statistics = json.loads(
        df.describe(include="all").to_json(date_format="iso")
    )

    return {
        "dataset_id": dataset.id,
        "file_name": dataset.file_name,
        "row_count": int(df.shape[0]),
        "column_count": int(df.shape[1]),
        "columns": [str(column) for column in df.columns],
        "preview": preview,
        "column_info": column_info,
        "duplicate_rows": duplicate_rows,
        "summary_statistics": summary_statistics,
    }


def detect_column_types(df: pd.DataFrame) -> dict[str, str]:
    column_types = {}

    for column in df.columns:
        if pd.api.types.is_bool_dtype(df[column]):
            column_types[column] = "boolean"
        elif pd.api.types.is_datetime64_any_dtype(df[column]):
            column_types[column] = "datetime"
        elif pd.api.types.is_numeric_dtype(df[column]):
            column_types[column] = "numerical"
        else:
            column_types[column] = "categorical"

    return column_types


def detect_missing_values(df: pd.DataFrame) -> dict[str, int]:
    missing_values = {}

    for column in df.columns:
        missing_values[column] = int(df[column].isna().sum())

    return missing_values


def detect_infinite_values(df: pd.DataFrame) -> dict[str, int]:
    infinite_values = {}

    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            count = int(df[column].isin([float("inf"), float("-inf")]).sum())
        else:
            count = 0
        infinite_values[column] = count

    return infinite_values


def detect_duplicates(df: pd.DataFrame) -> int:
    return int(df.duplicated().sum())


def build_cleaning_issues(
    column_types: dict[str, str],
    missing_values: dict[str, int],
    infinite_values: dict[str, int],
    duplicate_rows: int,
) -> list[dict]:
    issues = []

    for column, missing_count in missing_values.items():
        if missing_count > 0:
            column_type = column_types[column]

            if column_type == "numerical":
                recommended_action = "Fill missing values with median"
            elif column_type == "categorical":
                recommended_action = "Fill missing values with mode"
            else:
                recommended_action = "Fill missing values with most common value"

            issues.append(
                {
                    "column": column,
                    "issue_type": "missing_values",
                    "details": f"{missing_count} missing value(s) found",
                    "recommended_action": recommended_action,
                }
            )

    for column, infinite_count in infinite_values.items():
        if infinite_count > 0:
            issues.append(
                {
                    "column": column,
                    "issue_type": "infinite_values",
                    "details": f"{infinite_count} infinite value(s) found",
                    "recommended_action": "Replace infinite values and fill with median",
                }
            )

    if duplicate_rows > 0:
        issues.append(
            {
                "column": None,
                "issue_type": "duplicate_rows",
                "details": f"{duplicate_rows} duplicate row(s) found",
                "recommended_action": "Remove duplicate rows",
            }
        )

    return issues


def get_cleaning_report(
    db: Session,
    dataset_id: int,
    current_user: User,
) -> dict:
    dataset = get_owned_dataset(db, dataset_id, current_user)

    if not Path(dataset.file_path).is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset file not found on server",
        )

    df = read_stored_dataset_file(dataset.file_path)

    column_types = detect_column_types(df)
    missing_values = detect_missing_values(df)
    infinite_values = detect_infinite_values(df)
    duplicate_rows = detect_duplicates(df)

    issues = build_cleaning_issues(
        column_types=column_types,
        missing_values=missing_values,
        infinite_values=infinite_values,
        duplicate_rows=duplicate_rows,
    )

    return {
        "dataset_id": dataset.id,
        "file_name": dataset.file_name,
        "row_count": int(df.shape[0]),
        "column_count": int(df.shape[1]),
        "column_types": column_types,
        "missing_values": missing_values,
        "infinite_values": infinite_values,
        "duplicate_rows": duplicate_rows,
        "issues": issues,
        "ready_for_ml": len(issues) == 0,
    }


def clean_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    original_duplicate_count = int(df.duplicated().sum())

    cleaned_df = df.drop_duplicates().copy()

    column_types = detect_column_types(cleaned_df)

    for column, column_type in column_types.items():
        if column_type == "numerical":
            cleaned_df[column] = cleaned_df[column].replace(
                [float("inf"), float("-inf")],
                float("nan"),
            )

    for column in cleaned_df.columns:
        missing_count = int(cleaned_df[column].isna().sum())

        if missing_count == 0:
            continue

        if column_types[column] == "numerical":
            fill_value = cleaned_df[column].median()
            if pd.isna(fill_value):
                fill_value = 0
        else:
            mode_values = cleaned_df[column].mode()
            fill_value = mode_values.iloc[0] if not mode_values.empty else "Unknown"

        cleaned_df[column] = cleaned_df[column].fillna(fill_value)

    return cleaned_df, original_duplicate_count


def save_cleaned_dataset_file(
    cleaned_df: pd.DataFrame,
    original_file_path: str,
) -> str:
    original_path = Path(original_file_path)
    cleaned_file_path = original_path.with_name(f"cleaned_{original_path.stem}.csv")

    cleaned_df.to_csv(cleaned_file_path, index=False)

    return str(cleaned_file_path)


def clean_dataset(
    db: Session,
    dataset_id: int,
    current_user: User,
) -> dict:
    dataset = get_owned_dataset(db, dataset_id, current_user)

    if not Path(dataset.file_path).is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset file not found on server",
        )

    df = read_stored_dataset_file(dataset.file_path)
    original_row_count = int(df.shape[0])

    cleaned_df, removed_duplicate_rows = clean_dataframe(df)

    cleaned_file_path = save_cleaned_dataset_file(
        cleaned_df=cleaned_df,
        original_file_path=dataset.file_path,
    )

    try:
        dataset.cleaned_file_path = cleaned_file_path
        db.commit()
        db.refresh(dataset)
    except Exception:
        db.rollback()
        Path(cleaned_file_path).unlink(missing_ok=True)
        raise

    return {
        "dataset_id": dataset.id,
        "original_row_count": original_row_count,
        "cleaned_row_count": int(cleaned_df.shape[0]),
        "removed_duplicate_rows": removed_duplicate_rows,
        "message": "Dataset cleaned successfully",
    }
